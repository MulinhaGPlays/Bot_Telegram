from openpyxl import workbook, load_workbook

Planilha = load_workbook("Planilhas/P1.xlsx")

aba_ativa = Planilha.active

for celula in aba_ativa["C"]:
    if celula.value == "Unidade":
        linha = celula.row
        aba_ativa.cell(row=linha, column=5).value = 1

for celula in aba_ativa["C"]:
    if celula.value == "Unidade":
        linha = celula.row
        aba_ativa.cell(row=linha, column=4).value = "Uso Único"

Planilha.save("Planilhas/P1_AlterOpenpyxl.xlsx")
