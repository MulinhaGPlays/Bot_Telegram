from openpyxl import workbook, load_workbook # Importando o Openpyxl e o Workbook do Openpyxl para manipulação de arquivos Excel

Planilha = load_workbook("Bots_Automacao/Automacao_Excel/Planilhas/P1.xlsx")

aba_ativa = Planilha.active

for celula in aba_ativa["C"]:
    if celula.value == "Unidade":
        linha = celula.row
        aba_ativa.cell(row=linha, column=5).value = 1

for celula in aba_ativa["C"]:
    if celula.value == "Unidade":
        linha = celula.row
        aba_ativa.cell(row=linha, column=4).value = "Uso Único"

Planilha.save("Bots_Automacao/Automacao_Excel/Planilhas/P1_AlterOpenpyxl.xlsx")
